from django.shortcuts import render
from django.http import JsonResponse
from django.db import connection
import os,Global,xlrd
from api.models import *
import csv,io,json
import openpyxl as op


# Create your views here.
def demo(request):
    print(request)
    print(request.POST)
    username = request.POST.get('username')
    password = request.POST.get('password')
    result = {"username": username, "password": password}
    return JsonResponse(result)


def jira_bug_upload(request):
    if request.method == 'GET':
        result = {'error_code': 10003}
        return JsonResponse(result)
    elif request.method == 'POST':
        result = {}
        # 接受xls文件并保存到指定位置，文件类型为xls
        obj = request.FILES.get('file')
        bussiness_line = request.POST.get('bussiness_line')
        fix_version = request.POST.get('fix_version')
        cursor = connection.cursor()
        if obj and bussiness_line:
            f = open(os.path.join(Global.path, 'magic_site/static/upload/upload_file.csv'), 'wb')
            for line in obj.chunks():
                f.write(line)
            f.close()
            try:
                # 查询数据库case_step中，所有case等于case_id的集合，赋值给list名称为connect_list
                sql_team_person = 'SELECT team, person FROM magic_team_base WHERE bussiness_line = "%s"' % bussiness_line
                cursor.execute(sql_team_person)
                connect_list = cursor.fetchall()
                for i in connect_list:
                    if i[0] == '前端':
                        list_forward = json.loads(i[1])['person']
                    elif i[0] == '后端':
                        list_back = json.loads(i[1])['person']
                    elif i[0] == '美术':
                        list_art = json.loads(i[1])['person']
                    elif i[0] == '脚本':
                        list_jiaoben = json.loads(i[1])['person']
                    elif i[0] == '其他':
                        list_other = json.loads(i[1])['person']
                    elif i[0] == '特殊人员':
                        list_special = json.loads(i[1])['person']
                data = read_csv_list(Global.path + 'magic_site/static/upload/upload_file.csv', '经办人', '修复的版本')
                bug_dict = read_bug_data(data, list_forward, list_back, list_art, list_jiaoben, list_other, list_special)
                data_bug_F = bug_dict['bug_data']
                out_bug_F = bug_dict['out_data']
                data_fix_F = read_fix_data(data, fix_version)
                # write_column_insert(Global.path + 'magic_site/static/result/Bug数量&修复状态.xlsx', data_bug_F)
                # write_insert_block(Global.path + 'magic_site/static/result/Bug数量&修复状态.xlsx', 2)
                # write_rows_insert(Global.path + 'magic_site/static/result/Bug数量&修复状态.xlsx', data_fix_F)
                result = {'error_code': 10000, "data_bug_F": data_bug_F, "out_bug_F":out_bug_F, "data_fix_F":data_fix_F}
            except Exception as e:
                print(e)
                result = {'error_code': 10015}
                return JsonResponse(result)
        else:
            result = {'error_code': 10001}
    return JsonResponse(result)


# 读取csv文件
def read_csv_list(file_path, manager, version):
    bug = open(file_path, 'r', encoding='utf-8')
    data_list = []
    with bug:
        reader = csv.reader(bug)
        for row in reader:
            if row[0] == '问题关键字':
                for i in range(len(row)):
                    if row[i] == manager:
                        manager = i
                    if row[i] == version:
                        version = i
            else:
                data_list.append(row)
    return manager, version, data_list


# 读取csv文件
def read_csv_list(file_path, manager, version):
    bug = open(file_path, 'r', encoding='utf-8')
    data_list = []
    with bug:
        reader = csv.reader(bug)
        for row in reader:
            if row[0] == '问题关键字':
                for i in range(len(row)):
                    if row[i] == manager:
                        manager = i
                    if row[i] == version:
                        version = i
            else:
                data_list.append(row)
    return manager, version, data_list


# 统计方法，比对已存在的list中的数据，符合要求的存入新的list中
def read_insert_tuple(reader, num, belong_list, name):
    flag = 0
    data = {}
    for row in reader:
        if row[num] in belong_list:
            flag = flag + 1
    data['name'] = name
    data['number'] = flag
    return data


# 统计方法，比对具体字符串的数据，符合要求的存入新的list中
def read_insert_bug(reader, num,  str_a, name):
    flag = 0
    data = {}
    for row in reader:
        if str(str_a) in row[num]:
            flag = flag + 1
    data['name'] = name
    data['number'] = flag
    return data


# 统计方法，比对出了某个特定字符串的数据，除了这些字符串的其他所有存入新的list中
def read_out_side_bug(reader, num,  name, out_list):
    flag = 0
    data = {}
    for row in reader:
        if row[num] not in out_list:
            flag = flag + 1
    data['name'] = name
    data['number'] = flag
    return data


# 读取指定index为经办人的数据，进行统计
def read_bug_data(data_tuple,list_forward, list_back, list_art, list_jiaoben, list_other, list_special):
    data_list = data_tuple[2]
    index = data_tuple[0]
    bug_data = []
    out_data = []
    tuple_bug_1 = read_insert_tuple(data_list, index, list_forward, '前端')
    tuple_bug_2 = read_insert_tuple(data_list, index, list_back, '后端')
    tuple_bug_3 = read_insert_tuple(data_list, index, list_art, '美术')
    tuple_bug_4 = read_insert_tuple(data_list, index, list_jiaoben, '脚本')
    tuple_bug_5 = read_insert_tuple(data_list, index, list_other, '其他')
    flag = 0
    for row in data_list:
        if row[index] in list_special:
            if row[2] == '人物/服装穿模' or row[2] == '场景穿模' or row[2] == '2D美术' or row[2] == 'Avatar系统-UI':
                tuple_bug_3['number'] = tuple_bug_3['number'] + 1
            elif row[2] == '剧情' or row[2] == '幕后故事' or row[2] == 'Avatar系统-策划':
                tuple_bug_4['number'] = tuple_bug_4['number'] + 1
            elif row[2] == '多语言-Multilingual':
                tuple_bug_5['number'] = tuple_bug_5['number'] + 1
            else:
                flag = flag + 1
                out_data.append(row)
                # print(row)
    bug_data.append(tuple_bug_1)
    bug_data.append(tuple_bug_2)
    bug_data.append(tuple_bug_3)
    bug_data.append(tuple_bug_4)
    bug_data.append(tuple_bug_5)
    result = {"bug_data": bug_data, "out_data": out_data}
    return result


# 读取index为修复版本的数据，进行统计
def read_fix_data(data_tuple, fix_version):
    data_list = data_tuple[2]
    index = data_tuple[1]
    # print(data_list)
    fix_data = []
    tuple_fix_1 = read_insert_bug(data_list, index,  fix_version, '解决')
    tuple_fix_2 = read_insert_bug(data_list, index, '挂起', '挂起')
    tuple_fix_3 = read_insert_bug(data_list, index, 'Backlog', 'Backlog')
    tuple_fix_4 = read_insert_bug(data_list, index, '持续观察', '持续观察')
    out_list = ['挂起', 'Backlog', '持续观察', fix_version]
    tuple_fix_5 = read_outside_bug(data_list, index, '下版本修改', out_list)
    fix_data.append(tuple_fix_1)
    fix_data.append(tuple_fix_2)
    fix_data.append(tuple_fix_3)
    fix_data.append(tuple_fix_4)
    fix_data.append(tuple_fix_5)
    # print(fix_data)
    # print(tuple_fix_5)
    return fix_data


# 统计方法，比对出了某个特定字符串的数据，除了这些字符串的其他所有存入新的list中
def read_outside_bug(reader, num, name, out_list):
    flag = 0
    data = {}
    for row in reader:
        flag_num = 0
        for str1 in out_list:
            if str1 in row[num]:
                pass
            else:
                flag_num += 1
        if flag_num == 0:
            flag = flag + 1
    data['name'] = name
    data['number'] = flag
    return data


# 写入csv
def write_csv(file_path, data_list):
    book = op.Workbook()  # 创建工作簿对象
    sheet = book['Sheet']  # 创建子表
    for i in range(len(data_list)):
        d = data_list[i]['name'], data_list[i]['number']
        sheet.append(d)  # 每次写入一行
    book.save(file_path)


# 空行
def write_insert_block(file_path, index):
    # 读取csv
    book = op.load_workbook(file_path)
    # 选择工作表
    sheet = book['Sheet']
    block_rows = sheet.max_row  # 获得行数
    for i in range(0, index):
        for j in range(0, 2):
            sheet.cell(row=block_rows + i + 1, column=j + 1, value='')
    book.save(file_path)


# 行追加
def write_rows_insert(file_path, data_list):
    # 读取csv
    book = op.load_workbook(file_path)
    # 选择工作表
    sheet = book['Sheet']
    start_rows = sheet.max_row  # 获得行数
    for i in range(0, len(data_list)):
        for j in range(0, 2):
            if j == 0:
                sheet.cell(row=start_rows + i + 1, column=j + 1, value=data_list[i]['name'])
            else:
                sheet.cell(row=start_rows + i + 1, column=j + 1, value=data_list[i]['number'])
    book.save(file_path)


# 竖追加
def write_column_insert(file_path, data_list):
    # 读取csv
    book = op.load_workbook(file_path)
    # 选择工作表
    sheet = book['Sheet']
    start_column = sheet.max_column  # 获得行数
    for i in range(0, len(data_list)):
        for j in range(0, 2):
            if j == 0:
                sheet.cell(row=i + 1, column=start_column + j + 1, value=data_list[i]['name'])
            else:
                sheet.cell(row=i + 1, column=start_column + j + 1, value=data_list[i]['number'])
    book.save(file_path)
